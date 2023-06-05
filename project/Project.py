from polygon import RESTClient
import config
import openpyxl

client = RESTClient(config.API_KEY)



ticker = "BRK.A"
info = client.get_aggs(ticker=ticker, multiplier=1, timespan="day", from_="2021-01-09", to="2023-05-31")

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers
sheet["A1"] = "Open"
sheet["B1"] = "Close"
sheet["C1"] = "Day Change"
sheet["D1"] = "Percent Change"
sheet["E1"] = "Positive Days"
sheet["F1"] = "Negative Days"

pos = 0
neg = 0

# Write data to the sheet
for row, agg in enumerate(info, start=2):
    sheet.cell(row=row, column=1).value = agg.open
    sheet.cell(row=row, column=2).value = agg.close
    sheet.cell(row=row, column=3).value = agg.close-agg.open

    if agg.open > agg.close:
       neg+=1
       sheet.cell(row=row, column=3)

    if agg.open < agg.close:
        pos+=1

sheet.cell(row=2, column=4).value = pos
sheet.cell(row=2, column=5).value = neg

# Save the workbook
filename = "stock_data.xlsx"  # Specify the filename for the Excel file
workbook.save(filename)
print(f"Data saved to {filename}")
